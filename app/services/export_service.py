"""
Export Service – CSV and Excel generation for invoices.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models.invoice import Invoice


EXPORT_COLUMNS = [
    ("Invoice #",        "invoice_number"),
    ("Vendor",           "vendor_name"),
    ("Invoice Date",     "invoice_date"),
    ("Due Date",         "due_date"),
    ("Amount",           "total_amount"),
    ("Currency",         "currency"),
    ("Status",           "status"),
    ("Risk Flag",        "risk_flag"),
    ("Risk Score",       "risk_score"),
    ("Category",         "category"),
    ("Uploaded At",      "upload_timestamp"),
    ("Days Since Invoice","days_since_invoice"),
    ("Overdue",          "is_overdue"),
]


def export_csv(invoices: list[Invoice]) -> io.BytesIO:
    """Return a CSV BytesIO object for the given invoices."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([col for col, _ in EXPORT_COLUMNS])
    for inv in invoices:
        d = inv.to_dict(full=False)
        writer.writerow([_fmt(d.get(field)) for _, field in EXPORT_COLUMNS])
    return io.BytesIO(output.getvalue().encode("utf-8-sig"))


def export_excel(invoices: list[Invoice]) -> io.BytesIO:
    """Return an Excel (.xlsx) BytesIO object for the given invoices."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # ── Header row styling ─────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D0D7E4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [col for col, _ in EXPORT_COLUMNS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────
    risk_colors = {
        "SAFE":      "C6EFCE",
        "MODERATE":  "FFEB9C",
        "HIGH RISK": "FFC7CE",
        "DUPLICATE": "E2EFDA",
    }
    status_colors = {
        "approved": "C6EFCE",
        "rejected": "FFC7CE",
        "pending":  "FFEB9C",
    }

    for row_idx, inv in enumerate(invoices, start=2):
        d = inv.to_dict(full=False)
        for col_idx, (_, field) in enumerate(EXPORT_COLUMNS, start=1):
            value = _fmt(d.get(field))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

            # Colour-code risk_flag column
            risk_col = next((i for i, (h, _) in enumerate(EXPORT_COLUMNS, 1) if h == "Risk Flag"), None)
            status_col = next((i for i, (h, _) in enumerate(EXPORT_COLUMNS, 1) if h == "Status"), None)
            if col_idx == risk_col:
                flag = d.get("risk_flag", "SAFE")
                color = risk_colors.get(flag, "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
            if col_idx == status_col:
                status = d.get("status", "pending")
                color = status_colors.get(status, "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)

        ws.row_dimensions[row_idx].height = 18

    # ── Column widths ─────────────────────────────────────────────────────
    col_widths = [14, 25, 14, 14, 12, 10, 12, 14, 12, 20, 20, 14, 10]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Auto filter ───────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    _write_summary(ws2, invoices)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _write_summary(ws, invoices: list[Invoice]):
    ws["A1"] = "InvoiceIQ – Export Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Total Invoices"
    ws["B3"] = len(invoices)
    ws["A4"] = "Total Amount"
    ws["B4"] = sum(inv.total_amount for inv in invoices)
    ws["A5"] = "Generated At"
    ws["B5"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    risk_counts = {}
    status_counts = {}
    for inv in invoices:
        risk_counts[inv.risk_flag] = risk_counts.get(inv.risk_flag, 0) + 1
        status_counts[inv.status] = status_counts.get(inv.status, 0) + 1

    ws["A7"] = "Risk Breakdown"
    ws["A7"].font = Font(bold=True)
    for i, (k, v) in enumerate(risk_counts.items(), start=8):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v

    start = 8 + len(risk_counts) + 1
    ws[f"A{start}"] = "Status Breakdown"
    ws[f"A{start}"].font = Font(bold=True)
    for i, (k, v) in enumerate(status_counts.items(), start=start + 1):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


def _fmt(value) -> str:
    """Format a value for CSV/Excel cells."""
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.2f}"
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value)
